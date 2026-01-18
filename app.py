import os
import sys
import uuid
import signal
import openpyxl
import csv
from flask import Flask, render_template, request, jsonify, send_file

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

app = Flask(__name__, 
            template_folder=resource_path('templates'),
            static_folder=resource_path('static'))

# Better logging for bundled apps
import logging
log_path = os.path.join(os.path.expanduser("~"), "ExcelMergePro.log")
logging.basicConfig(filename=log_path, level=logging.DEBUG, 
                    format='%(asctime)s %(levelname)s: %(message)s')

UPLOAD_FOLDER = os.path.join(os.path.expanduser("~"), "ExcelMergeUploads")
try:
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    logging.info(f"Upload folder initialized at: {UPLOAD_FOLDER}")
except Exception as e:
    logging.error(f"Failed to create upload folder: {e}")

def get_excel_sheets(file_path):
    if file_path.endswith('.csv'):
        return ['CSV Default']
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        return wb.sheetnames
    except Exception as e:
        return [str(e)]

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files['file']
    ext = os.path.splitext(file.filename)[1].lower()
    temp_filename = f"{uuid.uuid4()}{ext}"
    temp_path = os.path.join(UPLOAD_FOLDER, temp_filename)
    file.save(temp_path)
    
    sheets = get_excel_sheets(temp_path)
    return jsonify({
        'filename': file.filename,
        'temp_path': temp_path,
        'sheets': sheets
    })

def read_csv_headers(file_path):
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        return next(reader, [])

def read_excel_headers(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return [str(cell) if cell is not None else "" for cell in row]
    return []

@app.route('/get_columns', methods=['POST'])
def get_columns():
    data = request.json
    try:
        # Load Target Headers
        if data['target_file'].lower().endswith('.csv'):
            t_cols = read_csv_headers(data['target_file'])
        else:
            t_cols = read_excel_headers(data['target_file'], data['target_sheet'])
        
        # Load Source Headers
        if data['source_file'].lower().endswith('.csv'):
            s_cols = read_csv_headers(data['source_file'])
        else:
            s_cols = read_excel_headers(data['source_file'], data['source_sheet'])
            
        return jsonify({
            'target_columns': t_cols,
            'source_columns': s_cols
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def get_column_data(file_path, sheet_name, col_name):
    """获取单列的所有数据"""
    data = []
    if file_path.lower().endswith('.csv'):
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                data.append(row.get(col_name))
    else:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        try:
            col_idx = headers.index(col_name) + 1
        except ValueError:
            return []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row[col_idx-1])
    return data

@app.route('/preview', methods=['POST'])
def preview():
    data = request.json
    try:
        t_keys = set(str(v).strip() for v in get_column_data(data['target_file'], data['target_sheet'], data['target_key']) if v is not None)
        s_keys = set(str(v).strip() for v in get_column_data(data['source_file'], data['source_sheet'], data['source_key']) if v is not None)
        
        matches = len(t_keys.intersection(s_keys))
        total = len(t_keys)
        
        return jsonify({
            'total': total,
            'matches': matches,
            'misses': total - matches
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/merge', methods=['POST'])
def merge():
    data = request.json
    try:
        # 为了合并，我们需要将 source 映射读入内存
        source_mapping = {} # { key: {col_name: value} }
        
        # 加载 Source 数据
        if data['source_file'].lower().endswith('.csv'):
            with open(data['source_file'], 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    key = str(row.get(data['source_key'])).strip()
                    if key:
                        source_mapping[key] = row
        else:
            wb_s = openpyxl.load_workbook(data['source_file'], data_only=True)
            ws_s = wb_s[data['source_sheet']]
            headers_s = [cell.value for cell in ws_s[1]]
            for row in ws_s.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers_s, row))
                key = str(row_dict.get(data['source_key'])).strip()
                if key:
                    source_mapping[key] = row_dict

        # 处理 Target 数据并写入新文件
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        
        if data['target_file'].lower().endswith('.csv'):
            with open(data['target_file'], 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                headers_t = next(reader)
                output_ws.append(headers_t)
                for row in reader:
                    row_dict = dict(zip(headers_t, row))
                    key = str(row_dict.get(data['target_key'])).strip()
                    if key in source_mapping:
                        for m in data['mapping']:
                            row_dict[m['tgt']] = source_mapping[key].get(m['src'])
                    output_ws.append([row_dict.get(h) for h in headers_t])
        else:
            wb_t = openpyxl.load_workbook(data['target_file'], data_only=True)
            ws_t = wb_t[data['target_sheet']]
            headers_t = [cell.value for cell in ws_t[1]]
            output_ws.append(headers_t)
            for row in ws_t.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers_t, row))
                key = str(row_dict.get(data['target_key'])).strip()
                if key in source_mapping:
                    for m in data['mapping']:
                        row_dict[m['tgt']] = source_mapping[key].get(m['src'])
                output_ws.append([row_dict.get(h) for h in headers_t])

        output_path = os.path.join(UPLOAD_FOLDER, f"result_{uuid.uuid4()}.xlsx")
        output_wb.save(output_path)
        
        return send_file(output_path, as_attachment=True)
    except Exception as e:
        logging.error(f"Merge error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/shutdown', methods=['POST'])
def shutdown():
    logging.info("Shutdown requested.")
    # Use os.kill to terminate the process group to ensure all threads die
    os.kill(os.getpid(), signal.SIGINT)
    return "Application shutting down..."

if __name__ == '__main__':

    import webbrowser
    from threading import Timer
    
    def open_browser():
        try:
            webbrowser.open("http://127.0.0.1:5001")
        except Exception as e:
            logging.error(f"Failed to open browser: {e}")

    Timer(1.5, open_browser).start()
    app.run(debug=False, port=5001)
